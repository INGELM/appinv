from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app.forms import ImportForm
from app import db
from app.models import Inventario
import tempfile
import os
from datetime import datetime
from io import BytesIO
import json

importar_bp = Blueprint('importar', __name__, url_prefix='/importar')


@importar_bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    if not current_user.tiene_permiso('ver_inventario'):
        flash('No tiene permiso para importar inventario', 'danger')
        return redirect(url_for('inventario.index'))

    form = ImportForm()
    if form.validate_on_submit():
        f = form.archivo.data
        # Import openpyxl lazily so app can start even if dependency is missing
        try:
            from openpyxl import load_workbook
        except ImportError:
            flash('La librería openpyxl no está instalada. Instale con: pip install openpyxl', 'danger')
            return redirect(url_for('inventario.index'))
        # Guardar temporalmente porque openpyxl en Windows necesita un path
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        tmp_name = tmp.name
        tmp.close()
        try:
            f.save(tmp_name)
            wb = load_workbook(tmp_name, data_only=True)
            ws = wb.active

            # Leer encabezados
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value is not None else '')

            # Columnas esperadas
            expected = ['material', 'cantidad', 'tipo']
            optional = ['num_proceso', 'proveedor', 'observaciones', 'descripcion']

            missing = [h for h in expected if h not in headers]
            if missing:
                flash(f'Archivo inválido. Faltan columnas: {", ".join(missing)}', 'danger')
                return redirect(url_for('importar.index'))

            col_index = {h: headers.index(h) for h in headers}

            # Parsear filas a estructura JSON para preview
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {'_fila': row_idx}
                try:
                    material = row[col_index.get('material')]
                    cantidad = row[col_index.get('cantidad')]
                    tipo = row[col_index.get('tipo')]

                    num_proceso = row[col_index.get('num_proceso')] if 'num_proceso' in col_index else None
                    proveedor = row[col_index.get('proveedor')] if 'proveedor' in col_index else None
                    observaciones = row[col_index.get('observaciones')] if 'observaciones' in col_index else None
                    descripcion = row[col_index.get('descripcion')] if 'descripcion' in col_index else None

                    if material is None or cantidad is None or tipo is None:
                        raise ValueError('Campos requeridos vacíos')

                    # Normalizar y validar
                    material_s = str(material).strip()
                    tipo_s = str(tipo).strip()
                    try:
                        cantidad_i = int(cantidad)
                    except Exception:
                        raise ValueError('Cantidad no es un entero')

                    if cantidad_i <= 0:
                        raise ValueError('Cantidad debe ser mayor que cero')

                    if tipo_s not in ['Herramienta', 'Consumible', 'Equipo', 'Repuesto', 'Accesorio']:
                        raise ValueError('Tipo inválido (usar Herramienta, Consumible, Equipo, Repuesto o Accesorio)')

                    row_data.update({
                        'material': material_s,
                        'cantidad': cantidad_i,
                        'tipo': tipo_s,
                        'descripcion': str(descripcion).strip() if descripcion else '',
                        'num_proceso': str(num_proceso).strip() if num_proceso else '',
                        'proveedor': str(proveedor).strip() if proveedor else '',
                        'observaciones': str(observaciones).strip() if observaciones else '',
                        'error': ''
                    })
                except Exception as e:
                    row_data.update({
                        'material': row[0] if row and len(row) > 0 else '',
                        'cantidad': row[1] if row and len(row) > 1 else '',
                        'tipo': row[2] if row and len(row) > 2 else '',
                        'descripcion': row[3] if row and len(row) > 3 else '',
                        'num_proceso': '',
                        'proveedor': '',
                        'observaciones': '',
                        'error': str(e)
                    })

                # Indicar si existe ya (por nombre exacto, case-insensitive)
                try:
                    existing = Inventario.query.filter(Inventario.material.ilike(row_data.get('material'))).first()
                except Exception:
                    existing = None
                row_data['exists'] = bool(existing)
                rows.append(row_data)

            # Render preview
            return render_template('inventario/importar_preview.html', rows=rows, form=form)
        finally:
            try:
                os.unlink(tmp_name)
            except Exception:
                pass

    return render_template('inventario/importar.html', form=form)


@importar_bp.route('/download_template', methods=['GET'])
@login_required
def download_template():
    # Genera dinámicamente un .xlsx con encabezados y filas de ejemplo
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('La librería openpyxl no está instalada. Instale con: pip install openpyxl', 'danger')
        return redirect(url_for('importar.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla'
    headers = ['material', 'cantidad', 'tipo', 'num_proceso', 'proveedor', 'observaciones', 'descripcion']
    ws.append(headers)
    # filas de ejemplo
    ws.append(['Tornillo M6 x 20', 100, 'Consumible', 'PROC-001', 'Proveedor A', 'Para uso general', 'Tornillo métricamente correcto'])
    ws.append(['Llave Inglesa 12mm', 5, 'Herramienta', 'PROC-002', 'Proveedor B', 'Herramienta compartida', '12mm, llave ajustable'])
    ws.append(['Filtro de Aceite', 10, 'Repuesto', 'PROC-003', 'Proveedor C', 'Repuesto para mantenimiento', 'Filtro modelo ABC'])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(bio,
                     as_attachment=True,
                     download_name='import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@importar_bp.route('/confirm', methods=['POST'])
@login_required
def confirm():
    if not current_user.tiene_permiso('ver_inventario'):
        flash('No tiene permiso para importar inventario', 'danger')
        return redirect(url_for('inventario.index'))

    # Esperamos: action_mode, rows_json, selected (opcional list of indexes)
    action_mode = request.form.get('action_mode', 'crear_nuevos')
    rows_json = request.form.get('rows_json')
    selected = request.form.getlist('selected')

    if not rows_json:
        flash('Datos de importación no encontrados', 'danger')
        return redirect(url_for('importar.index'))

    try:
        rows = json.loads(rows_json)
    except Exception:
        flash('Formato de datos inválido', 'danger')
        return redirect(url_for('importar.index'))

    # Si no hay selección, tomar todas
    if not selected:
        indices = list(range(len(rows)))
    else:
        indices = [int(i) for i in selected]

    inserted = 0
    updated = 0
    errors = []
    inserted_items = []
    updated_items = []

    for idx in indices:
        try:
            r = rows[idx]
            if r.get('error'):
                errors.append(f"Fila {r.get('_fila')}: {r.get('error')}")
                continue

            material = r['material']
            cantidad = int(r['cantidad'])
            tipo = r['tipo']
            num_proceso = r.get('num_proceso') or None
            proveedor = r.get('proveedor') or None
            observaciones = r.get('observaciones') or None
            descripcion = r.get('descripcion') or None

            existing = Inventario.query.filter(Inventario.material.ilike(material)).first()

            if existing:
                if action_mode == 'crear_nuevos':
                    # crear duplicado con mismo nombre
                    inv = Inventario(
                        fecha_ingreso=datetime.utcnow(),
                        num_proceso=num_proceso,
                        proveedor=proveedor,
                        material=material,
                        descripcion=descripcion,
                        cantidad=cantidad,
                        tipo=tipo,
                        observaciones=observaciones
                    )
                    db.session.add(inv)
                    inserted += 1
                    inserted_items.append(f"{material} ({cantidad})")
                elif action_mode == 'actualizar_existente':
                    # actualizar campos existentes (sumar cantidad)
                    prev_cant = existing.cantidad
                    existing.cantidad = existing.cantidad + cantidad
                    # actualizar otros campos si vienen
                    if num_proceso:
                        existing.num_proceso = num_proceso
                    if proveedor:
                        existing.proveedor = proveedor
                    if observaciones:
                        # concatenar observaciones
                        existing.observaciones = (existing.observaciones or '') + '\n' + observaciones
                    if descripcion:
                        existing.descripcion = descripcion
                    existing.tipo = tipo
                    updated += 1
                    updated_items.append(f"{material} (antes {prev_cant}, ahora {existing.cantidad})")
                else:  # crear_o_actualizar
                    prev_cant = existing.cantidad
                    existing.cantidad = existing.cantidad + cantidad
                    if num_proceso:
                        existing.num_proceso = num_proceso
                    if proveedor:
                        existing.proveedor = proveedor
                    if observaciones:
                        existing.observaciones = (existing.observaciones or '') + '\n' + observaciones
                    if descripcion:
                        existing.descripcion = descripcion
                    existing.tipo = tipo
                    updated += 1
                    updated_items.append(f"{material} (antes {prev_cant}, ahora {existing.cantidad})")
            else:
                # No existe -> crear
                inv = Inventario(
                    fecha_ingreso=datetime.utcnow(),
                    num_proceso=num_proceso,
                    proveedor=proveedor,
                    material=material,
                    descripcion=descripcion,
                    cantidad=cantidad,
                    tipo=tipo,
                    observaciones=observaciones
                )
                db.session.add(inv)
                inserted += 1
                inserted_items.append(f"{material} ({cantidad})")
        except Exception as e:
            errors.append(f"Fila {r.get('_fila') if 'r' in locals() else idx}: {str(e)}")

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error al guardar en la base de datos: {str(e)}', 'danger')
        return redirect(url_for('importar.index'))

    msg = f'Importación procesada. Insertados: {inserted}. Actualizados: {updated}. Errores: {len(errors)}.'
    if errors:
        msg += ' Revise los detalles en los mensajes.'
        for err in errors:
            flash(err, 'warning')

    # Mensajes con detalle de qué se insertó/actualizó
    if inserted_items:
        flash('Insertados: ' + ', '.join(inserted_items), 'info')
    if updated_items:
        flash('Actualizados: ' + ', '.join(updated_items), 'info')

    flash(msg, 'success' if (inserted + updated) > 0 else 'warning')
    return redirect(url_for('inventario.index'))
